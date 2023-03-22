import concurrent.futures
import os
import re
import string
import subprocess
import sys
from time import sleep

import yt_dlp
from openpyxl import load_workbook

# --------- Adjustable options for ffmpeg---------
ffmpeg_opts = "-crf 28 -b:a 256 -preset slow -filter:v fps=30 -codec:v libx264 -n "
# --------- Adjustable options for yt_dlp---------
yt_dlp_opts = "--format \"bv*+ba/b\" --write-thumbnail "

links = []
titles = []
descriptions = []



def linksFromExcel():
    # Selecting specific sheet
    wb = load_workbook("input.xlsx")
    sheet = wb.active
    col = sheet["A"]
    for cell in col:
        if isinstance(cell.value, str):
            links.append(cell.value)

    # for data in links:
    #     print(data)


def writeToExcel():
    wb = load_workbook("output.xlsx")
    sheet = wb.active

    # reading specific column
    sheet.cell(row=1, column=1).value = "Title"
    sheet.cell(row=1, column=2).value = "Description"
    sheet.cell(row=1, column=3).value = "Links"
    i = 2

    while i - 2 < len(links):
        sheet.cell(row=i, column=1).value = titles[i - 2]
        sheet.cell(row=i, column=2).value = descriptions[i - 2]
        sheet.cell(row=i, column=3).value = links[i - 2]
        i += 1

    wb.save("output.xlsx")


def download(path, root):
    linksFromExcel()

    os.chdir(path)
    print(len(links))
    for link in range(len(links)):
        # --start of---- Download information about the video-------
        with yt_dlp.YoutubeDL() as ydl:
            info = ydl.extract_info(links[link], download=False)
            jsonobj = ydl.sanitize_info(info)
        description = jsonobj['description']
        descriptions.append(description)
        title = jsonobj['title']

        # --end of---- Download information about the video-------
        # -
        # -
        # --start of----- Creating directory according to title of video------
        titles.append(title)

        if not title.isalnum():
            pattern = r'[' + string.punctuation + ']'
            title = re.sub(pattern, '', title)

        if not os.path.exists(os.getcwd() + "\\Downloaded_Videos"):
            os.mkdir("Downloaded_Videos")
        # print(os.getcwd())
        os.chdir("Downloaded_Videos")
        if not os.path.exists(os.getcwd() + "\\" + title):
            try:
                os.mkdir(title)

            except Exception:
                print("Invalid Folder Name Cannot use video title as Folder Name!")

        os.chdir(title)

        # --end of----- Creating directory according to title of video------
        # -
        # -
        # --start of ----- Downloading video------
        file_name = title
        if subprocess.run(f'yt-dlp {yt_dlp_opts} -o \"{file_name}\".%(ext)s {links[link]}', shell=True).returncode:
            # with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            #     if ydl.download(links[link]):
            print("Error downloading file")
        # --end of ----- Downloading video------
        # -
        # -
        # --start of---extracting filename and using it to extract thumbnail --
        available_files = subprocess.run('dir /b', shell=True, capture_output=True).stdout.decode()
        # print(available_files)
        # print("title " + title.split(" ")[0] + "end")
        # print(available_files.find(file_name))
        # file_extension = available_files[available_files.find(file_name):len(file_name) + 5].split(".")[1].strip()

        # filename.replace("|", "\|")
        # print("filename is " + f'{file_name}.{file_extension}' + "ends here")
        # print(f'{file_name}.{file_extension}')
        if subprocess.run(f"ffmpeg -y -i \"{file_name}.webp\"  thumbnail.jpg", shell=True
                , capture_output=True).returncode:
            print(f"Attention! Video \"{file_name}\" thumbnail might have not been generated. ")

        os.remove(f"{file_name}.webp")
        # print(str(thumbNail.stdout.decode().strip()))
        # --start of---extracting filename and using it to extract thumbnail --
        # -
        # -
        # --start of--- writing info file to title folder---
        file = open("info.txt", "w", encoding="utf-8")

        file.write("Tile: " + str(title) + "\n" + "description: " + str(description))
        # --end of--- writing file to title folder---

        os.chdir("..")
        os.chdir("..")

    os.chdir(root)
    writeToExcel()
    # print("len of folenames " + str(len(filenames)))
    return 'Download Process completed'


# results = []
def compress(path):
    visited_folders_list = []
    folders_list = []
    os.chdir(path)
    print("P2 in " + os.getcwd())
    while True:
        folders = subprocess.run("dir /b /od Downloaded_Videos", capture_output=True, shell=True).stdout.decode()

        folders_list = []

        for x in range(len(folders.split("\n")) - 1):
            folders_list.append(folders.split("\n")[x])
        # print(folders_list)

        while len(folders_list) < 2:
            sleep(5)
            folders = subprocess.run("dir /b /od Downloaded_Videos", capture_output=True, shell=True).stdout.decode()

            folders_list = []

            for x in range(len(folders.split("\n")) - 1):
                folders_list.append(folders.split("\n")[x])
            # print(folders_list)
        remove = []
        for x in range(0, len(visited_folders_list)):
            for y in range(0, len(folders_list)):
                # print("folders list "+folders_list[y])
                # print("visited list "+visited_folders_list[x])
                if folders_list[y] == visited_folders_list[x]:
                    remove.append(x)
        for x in remove:
            if visited_folders_list[x] in folders_list:
                folders_list.remove(visited_folders_list[x])

        # print("len is " + str(len(folders_list)))
        if len(folders_list) == 0:
            break
            # print(folders_list)

        i = 0
        while i < len(folders_list):
            files = subprocess.run(f"dir /b /od Downloaded_Videos\\\"{folders_list[i]}\"", capture_output=True,
                                   shell=True).stdout.decode().split("\n")
            visited_folders_list.append(folders_list[i])
            # print(files)
            if len(files) < 5:
                print(len(files))

                for file in files:
                    if file.split(" ")[0] == folders_list[i].split(" ")[0]:

                        print(f"Starting Compression for {file}")

                        if subprocess.run(
                                    f"ffmpeg -i \"Downloaded_Videos\\{folders_list[i]}\\{file}\" {ffmpeg_opts} \"Downloaded_Videos\\{folders_list[i]}\\{file.split('.')[0]}_converted.mp4\"",
                                    shell=True, capture_output=True
                            ).returncode:
                                print(
                                    "Attention! Video Conversion Might have errors. or its already done if its corrupt delete the file \n and re run the script. if its already converted ignore this message \n")
                        else:
                                # os.remove(f'\"Downloaded_Videos\\{folders_list[i]}\\{file}\"')
                                print("Video has been compresses Successfully  \n")
                else:
                  print(f"{file} already compressed")

            i = i + 1

    return 'Compressions Process completed '


def main():
    cwd = os.getcwd()
    path = os.getcwd()

    if len(sys.argv) == 2 and os.path.exists(sys.argv[1]):
        path = sys.argv[1]
    with concurrent.futures.ProcessPoolExecutor() as executor:
        p1 = executor.submit(download, path, cwd)
        p2 = executor.submit(compress, path)
        print(p1.result())
        print(p2.result())


if __name__ == '__main__':
    main()
